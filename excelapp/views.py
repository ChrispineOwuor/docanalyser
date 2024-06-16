from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import workbook, load_workbook
from excelapp.serializers import FileSerializer


# Create your views here.
class AnalyseFile(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request):

        return Response("success", status=status.HTTP_200_OK)

    def post(self, request):

        try:
            file_serializer = FileSerializer(data=request.data)
            if file_serializer.is_valid():
                doc = file_serializer.validated_data['filename']
                book = load_workbook(doc)
                sheet = book.active
                for row in sheet.iter_rows(values_only=True):
                    if 'Total Assets' in row:
                        TotalAssets = row[1]
                    elif 'Current assets' in row:
                        Currentassets = row[1]
                    elif 'Current liabilities' in row:
                        Currentliabilities = row[1]
                    elif 'Total Liabilities' in row:
                        TotalLiabilities = row[1]
                CurrentRatio = Currentassets / Currentliabilities
                DebtRatio = TotalLiabilities / TotalAssets

                res = {
                    "TotalAssets": TotalAssets,
                    "Currentassets": Currentassets,
                    "Currentliabilities": Currentliabilities,
                    "TotalLiabilities": TotalLiabilities,
                    "CurrentRatio": CurrentRatio,
                    "DebtRatio": DebtRatio
                }

                return Response({"file": res}, status=status.HTTP_200_OK)
            return Response(file_serializer.errors, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            return Response(str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR)
